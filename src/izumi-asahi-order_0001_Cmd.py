"""Excelファイル内の「BY」を含むシートをTSV出力するコマンドラインツール。"""

from __future__ import annotations

import csv
import sys
from pathlib import Path
from typing import Iterable


START_MESSAGE = "izumi-asahi-order_0001 started"
TARGET_SHEET_KEYWORD = "BY"
USAGE = "Usage: python src/izumi-asahi-order_0001_Cmd.py <excel_file_path>"


def build_output_path(excel_path: Path, sheet_index: int) -> Path:
    """BYシートの順番に応じたTSV出力パスを作成する。"""
    if sheet_index == 1:
        return excel_path.with_suffix(".tsv")

    return excel_path.with_name(f"{excel_path.stem}_{sheet_index:04d}.tsv")


def normalize_cell_value(value: object) -> str:
    """TSV出力用にセル値を文字列へ変換する。"""
    if value is None:
        return ""

    return str(value)


def write_sheet_to_tsv(rows: Iterable[tuple[object, ...]], output_path: Path) -> None:
    """Excelシートの行データをTSVファイルへ出力する。"""
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.writer(output_file, delimiter="\t", lineterminator="\n")
        for row in rows:
            writer.writerow([normalize_cell_value(cell) for cell in row])


def export_by_sheets_to_tsv(excel_path: Path) -> int:
    """シート名にBYを含むシートだけをTSVファイルとして出力する。"""
    from openpyxl import load_workbook

    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    by_sheet_names = [
        sheet_name
        for sheet_name in workbook.sheetnames
        if TARGET_SHEET_KEYWORD in sheet_name
    ]

    if not by_sheet_names:
        print("No sheets containing 'BY' were found.")
        workbook.close()
        return 0

    for index, sheet_name in enumerate(by_sheet_names, start=1):
        worksheet = workbook[sheet_name]
        output_path = build_output_path(excel_path, index)
        write_sheet_to_tsv(worksheet.iter_rows(values_only=True), output_path)
        print(f"Exported sheet '{sheet_name}' to '{output_path}'")

    workbook.close()
    return len(by_sheet_names)


def main() -> int:
    """コマンドライン引数を読み取り、ExcelからTSVへの出力を実行する。"""
    print(START_MESSAGE)

    if len(sys.argv) < 2:
        print(USAGE)
        return 1

    excel_path = Path(sys.argv[1])

    if not excel_path.exists():
        print(f"Error: file not found: {excel_path}")
        return 1

    if not excel_path.is_file():
        print(f"Error: not a file: {excel_path}")
        return 1

    if excel_path.suffix.lower() != ".xlsx":
        print(f"Error: expected an .xlsx file: {excel_path}")
        return 1

    export_by_sheets_to_tsv(excel_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
