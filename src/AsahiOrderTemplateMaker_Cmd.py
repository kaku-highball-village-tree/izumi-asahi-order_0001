# -- coding: utf-8 --
###############################################################
#
# AsahiOrderTemplateMaker_Cmd.py
#
# pip install openpyxl tkcalendar
#
###############################################################

import csv
import os
import shutil
import sys
import tempfile
import tkinter as tk
from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import messagebox

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from tkcalendar import Calendar


EXPECTED_HEADERS: tuple[str, str, str] = ("productCode", "productName", "spec")
STEP0002_HEADERS: tuple[str, ...] = (
    "納品日",
    "曜日",
    "配送パターン",
    "便",
    "Ｐ品番",
    "APEX品番",
    "商品名",
    "産地",
    "仕様",
    "伝票原価",
    "伝票売価",
    "値入",
    "売価",
    "単位",
)
WEEKDAYS: tuple[str, ...] = ("月", "火", "水", "木", "金", "土", "日")
SUPPORTED_EXTENSIONS: set[str] = {".xlsx", ".tsv", ".csv"}


class ProductRow:
    """テンプレートへ出力する1商品の3列を保持します。"""

    def __init__(self, product_code: str, product_name: str, spec: str) -> None:
        self.product_code: str = product_code
        self.product_name: str = product_name
        self.spec: str = spec


class Step0002Error(Exception):
    """処理0002で発生したエラーであることを呼び出し元へ伝えます。"""


class Step0003Error(Exception):
    """処理0003で発生したエラーであることを呼び出し元へ伝えます。"""


class Step0004Error(Exception):
    """処理0004で発生したエラーであることを呼び出し元へ伝えます。"""


def write_error_text(pszOutputFileFullPath: str, pszErrorMessage: str) -> None:
    """エラーメッセージをUTF-8のテキストファイルへ上書き保存します。"""
    pszDirectoryFullPath: str = os.path.dirname(pszOutputFileFullPath)
    if pszDirectoryFullPath != "":
        os.makedirs(pszDirectoryFullPath, exist_ok=True)
    with open(pszOutputFileFullPath, mode="w", encoding="utf-8", newline="") as objFile:
        objFile.write(pszErrorMessage.rstrip("\n") + "\n")


def get_error_file_full_path(pszInputFileFullPath: str) -> str:
    """入力ファイルと同じフォルダーに作る_error.txtのパスを返します。"""
    pszDirectoryFullPath: str = os.path.dirname(os.path.abspath(pszInputFileFullPath))
    pszBaseNameWithoutExtension: str = os.path.splitext(
        os.path.basename(pszInputFileFullPath)
    )[0]
    return os.path.join(pszDirectoryFullPath, pszBaseNameWithoutExtension + "_error.txt")


def report_processing_error(
    pszInputFileFullPath: str, pszProcessName: str, pszDetailMessage: str
) -> None:
    """標準エラーと入力ファイル用_error.txtへ同じエラーを出力します。"""
    pszErrorMessage: str = (
        "処理結果: エラー\n"
        + "入力ファイル: "
        + os.path.abspath(pszInputFileFullPath)
        + "\n発生した処理: "
        + pszProcessName
        + "\nエラー内容: "
        + pszDetailMessage
        + "\n"
    )
    print(pszErrorMessage, file=sys.stderr, end="")
    try:
        write_error_text(get_error_file_full_path(pszInputFileFullPath), pszErrorMessage)
    except OSError as objException:
        print(
            "Error: _error.txtの保存にも失敗しました。Detail = " + str(objException),
            file=sys.stderr,
        )


def remove_old_error_file(pszInputFileFullPath: str) -> None:
    """正常終了後、以前の処理で作られた_error.txtがあれば削除します。"""
    pszErrorFileFullPath: str = get_error_file_full_path(pszInputFileFullPath)
    if os.path.exists(pszErrorFileFullPath):
        os.remove(pszErrorFileFullPath)


def validate_input_path(pszInputFileFullPath: str) -> str:
    """入力パスと拡張子を検証し、絶対パスを返します。"""
    pszAbsolutePath: str = os.path.abspath(pszInputFileFullPath)
    if not os.path.exists(pszAbsolutePath):
        raise ValueError("入力ファイルが見つかりません。Path = " + pszAbsolutePath)
    if not os.path.isfile(pszAbsolutePath):
        raise ValueError("入力パスがファイルではありません。Path = " + pszAbsolutePath)
    pszExtension: str = os.path.splitext(pszAbsolutePath)[1].lower()
    if pszExtension not in SUPPORTED_EXTENSIONS:
        raise ValueError("入力ファイルの拡張子は未対応です。Path = " + pszAbsolutePath)
    return pszAbsolutePath


def normalize_header(objValue: object) -> str:
    """ヘッダー比較用に値を前後空白のない文字列へ変換します。"""
    if objValue is None:
        return ""
    return str(objValue).strip()


def normalize_text(objValue: object) -> str:
    """出力用にNoneを空文字、それ以外を文字列へ変換します。"""
    if objValue is None:
        return ""
    return str(objValue)


def get_this_week_monday(objToday: date | None = None) -> date:
    """基準日を含む週の月曜日を返します。"""
    objBaseDate: date = date.today() if objToday is None else objToday
    return objBaseDate - timedelta(days=objBaseDate.weekday())


def validate_start_monday(objStartMonday: date) -> None:
    """開始日が月曜日であることを確認します。"""
    if objStartMonday.weekday() != 0:
        raise ValueError(
            "開始日は月曜日を指定してください。Date = "
            + objStartMonday.isoformat()
        )


def parse_start_monday(pszValue: str) -> date:
    """YYYY-MM-DDを開始月曜日として解析・検証します。"""
    if len(pszValue) != 10:
        raise ValueError("開始日の形式はYYYY-MM-DDではありません。Value = " + pszValue)
    try:
        objStartMonday: date = date.fromisoformat(pszValue)
    except ValueError as objException:
        raise ValueError(
            "開始日の形式はYYYY-MM-DDではありません。Value = " + pszValue
        ) from objException
    if objStartMonday.isoformat() != pszValue:
        raise ValueError("開始日の形式はYYYY-MM-DDではありません。Value = " + pszValue)
    validate_start_monday(objStartMonday)
    return objStartMonday


def select_start_monday() -> date | None:
    """カレンダーを表示し、利用者が選んだ月曜日またはキャンセル時Noneを返します。"""
    objThisWeekMonday: date = get_this_week_monday()
    objNextWeekMonday: date = objThisWeekMonday + timedelta(days=7)
    objSelectedMonday: date | None = objNextWeekMonday
    objLastValidMonday: date = objNextWeekMonday

    objRoot = tk.Tk()
    objRoot.title("Asahi Order Template Maker - 開始月曜日の選択")
    objRoot.resizable(False, False)

    objInstructionLabel = tk.Label(
        objRoot,
        text="処理0004の開始月曜日を選択してください。",
        padx=10,
        pady=8,
    )
    objInstructionLabel.pack()

    objCalendar = Calendar(
        objRoot,
        selectmode="day",
        year=objNextWeekMonday.year,
        month=objNextWeekMonday.month,
        day=objNextWeekMonday.day,
        date_pattern="yyyy-mm-dd",
        firstweekday="monday",
    )
    objCalendar.pack(padx=10, pady=5)

    def set_selected_monday(objMonday: date) -> None:
        nonlocal objSelectedMonday, objLastValidMonday
        objSelectedMonday = objMonday
        objLastValidMonday = objMonday
        objCalendar.selection_set(objMonday)

    def on_calendar_selected(_objEvent: object = None) -> None:
        nonlocal objSelectedMonday
        objSelectedDate: date = objCalendar.selection_get()
        if objSelectedDate.weekday() != 0:
            objCalendar.selection_set(objLastValidMonday)
            messagebox.showerror(
                "Asahi Order Template Maker",
                "月曜日を選択してください。",
                parent=objRoot,
            )
            return
        objSelectedMonday = objSelectedDate

    def confirm_selection() -> None:
        nonlocal objSelectedMonday
        objSelectedDate: date = objCalendar.selection_get()
        if objSelectedDate.weekday() != 0:
            messagebox.showerror(
                "Asahi Order Template Maker",
                "月曜日を選択してください。",
                parent=objRoot,
            )
            return
        objSelectedMonday = objSelectedDate
        objRoot.destroy()

    def cancel_selection() -> None:
        nonlocal objSelectedMonday
        objSelectedMonday = None
        objRoot.destroy()

    objCalendar.bind("<<CalendarSelected>>", on_calendar_selected)
    objButtonFrame = tk.Frame(objRoot, padx=10, pady=10)
    objButtonFrame.pack(fill=tk.X)
    tk.Button(
        objButtonFrame,
        text="今週の月曜日",
        command=lambda: set_selected_monday(objThisWeekMonday),
    ).pack(side=tk.LEFT, padx=2)
    tk.Button(
        objButtonFrame,
        text="来週の月曜日",
        command=lambda: set_selected_monday(objNextWeekMonday),
    ).pack(side=tk.LEFT, padx=2)
    tk.Button(objButtonFrame, text="決定", command=confirm_selection).pack(
        side=tk.LEFT, padx=8
    )
    tk.Button(objButtonFrame, text="キャンセル", command=cancel_selection).pack(
        side=tk.LEFT, padx=2
    )
    objRoot.protocol("WM_DELETE_WINDOW", cancel_selection)
    objRoot.bind("<Escape>", lambda _objEvent: cancel_selection())
    objRoot.lift()
    objRoot.attributes("-topmost", True)
    objRoot.after_idle(lambda: objRoot.attributes("-topmost", False))
    objRoot.mainloop()
    return objSelectedMonday


def show_start_monday_cancelled_message() -> None:
    """開始月曜日の選択キャンセルにより処理を中止することを通知します。"""
    objRoot = tk.Tk()
    objRoot.withdraw()
    messagebox.showerror(
        "Asahi Order Template Maker",
        "開始月曜日の選択がキャンセルされました。\n"
        "処理0004を完了できないため、処理を中止します。",
        parent=objRoot,
    )
    objRoot.destroy()


def validate_headers(listValues: list[object], pszSourceName: str) -> None:
    """先頭3列が仕様どおりのヘッダーであることを確認します。"""
    if len(listValues) < 3:
        raise ValueError(pszSourceName + "のヘッダーは3列未満です。")
    tupleHeaders: tuple[str, str, str] = tuple(
        normalize_header(objValue) for objValue in listValues[:3]
    )  # type: ignore[assignment]
    if tupleHeaders != EXPECTED_HEADERS:
        raise ValueError(
            pszSourceName
            + "の先頭3列はproductCode、productName、specではありません。"
        )


def find_target_worksheet(objWorkbook: Workbook) -> Worksheet:
    """A1～C1が仕様どおりのシートを探し、1枚だけなら返します。"""
    listTargetWorksheets: list[Worksheet] = []
    for objWorksheet in objWorkbook.worksheets:
        listHeaders: list[object] = [
            objWorksheet.cell(row=1, column=iColumn).value for iColumn in range(1, 4)
        ]
        if tuple(normalize_header(objValue) for objValue in listHeaders) == EXPECTED_HEADERS:
            listTargetWorksheets.append(objWorksheet)
    if len(listTargetWorksheets) == 0:
        raise ValueError(
            "A1～C1がproductCode、productName、specのシートが見つかりません。"
        )
    if len(listTargetWorksheets) > 1:
        raise ValueError(
            "対象シートが複数見つかりました。対象シート = "
            + ", ".join(objWorksheet.title for objWorksheet in listTargetWorksheets)
        )
    return listTargetWorksheets[0]


def build_product_row(listValues: list[object], iRow: int) -> ProductRow | None:
    """先頭3列を検証し、空行ならNone、それ以外なら商品行を返します。"""
    if len(listValues) < 3:
        if all(normalize_text(objValue).strip() == "" for objValue in listValues):
            return None
        raise ValueError(str(iRow) + "行目は3列未満です。")
    listTexts: list[str] = [normalize_text(objValue) for objValue in listValues[:3]]
    if all(pszValue.strip() == "" for pszValue in listTexts):
        return None
    pszProductCode: str = listTexts[0].strip()
    if pszProductCode == "":
        raise ValueError(
            str(iRow) + "行目はproductCodeが空ですが、ほかの対象列に値があります。"
        )
    return ProductRow(pszProductCode, listTexts[1], listTexts[2])


def validate_unique_product_codes(
    listProductRows: list[ProductRow], pszSourceName: str
) -> None:
    """productCodeが重複していないことを確認します。"""
    setSeenCodes: set[str] = set()
    setDuplicateCodes: set[str] = set()
    for objProductRow in listProductRows:
        if objProductRow.product_code in setSeenCodes:
            setDuplicateCodes.add(objProductRow.product_code)
        setSeenCodes.add(objProductRow.product_code)
    if setDuplicateCodes:
        raise ValueError(
            pszSourceName
            + "でproductCodeが重複しています。productCode = "
            + ", ".join(sorted(setDuplicateCodes))
        )


def read_excel_rows(pszInputFileFullPath: str) -> list[ProductRow]:
    """Excelの対象シートからテンプレート用商品行を読み取ります。"""
    objWorkbook: Workbook = load_workbook(pszInputFileFullPath, data_only=True)
    objWorksheet: Worksheet = find_target_worksheet(objWorkbook)
    listProductRows: list[ProductRow] = []
    for iRow in range(2, objWorksheet.max_row + 1):
        listValues: list[object] = [
            objWorksheet.cell(row=iRow, column=iColumn).value for iColumn in range(1, 4)
        ]
        objProductRow: ProductRow | None = build_product_row(listValues, iRow)
        if objProductRow is not None:
            listProductRows.append(objProductRow)
    validate_unique_product_codes(listProductRows, "Excel")
    return listProductRows


def read_delimited_rows_with_encoding(
    pszInputFileFullPath: str, pszEncoding: str, pszDelimiter: str
) -> list[list[str]]:
    """指定文字コードと区切り文字で全レコードを読み取ります。"""
    with open(
        pszInputFileFullPath, mode="r", encoding=pszEncoding, newline=""
    ) as objFile:
        return list(csv.reader(objFile, delimiter=pszDelimiter, strict=True))


def read_delimited_rows(pszInputFileFullPath: str) -> list[ProductRow]:
    """CSVまたはTSVをUTF-8優先、失敗時CP932で読み取ります。"""
    pszExtension: str = os.path.splitext(pszInputFileFullPath)[1].lower()
    pszDelimiter: str = "\t" if pszExtension == ".tsv" else ","
    pszSourceName: str = "TSV" if pszExtension == ".tsv" else "CSV"
    try:
        listRows: list[list[str]] = read_delimited_rows_with_encoding(
            pszInputFileFullPath, "utf-8-sig", pszDelimiter
        )
    except UnicodeDecodeError:
        listRows = read_delimited_rows_with_encoding(
            pszInputFileFullPath, "cp932", pszDelimiter
        )
    if not listRows:
        raise ValueError(pszSourceName + "ファイルが空です。")
    validate_headers(listRows[0], pszSourceName)
    listProductRows: list[ProductRow] = []
    for iRow, listValues in enumerate(listRows[1:], start=2):
        objProductRow: ProductRow | None = build_product_row(listValues, iRow)
        if objProductRow is not None:
            listProductRows.append(objProductRow)
    validate_unique_product_codes(listProductRows, pszSourceName)
    return listProductRows


def get_output_file_paths(pszInputFileFullPath: str) -> tuple[Path, Path]:
    """_step0001.xlsxと_step0001.tsvの出力パスを返します。"""
    objInputPath: Path = Path(pszInputFileFullPath)
    objBasePath: Path = objInputPath.with_name(objInputPath.stem + "_step0001")
    return objBasePath.with_suffix(".xlsx"), objBasePath.with_suffix(".tsv")


def get_step0002_output_file_paths(pszInputFileFullPath: str) -> tuple[Path, Path]:
    """_step0002.xlsxと_step0002.tsvの出力パスを返します。"""
    objInputPath: Path = Path(pszInputFileFullPath)
    objBasePath: Path = objInputPath.with_name(objInputPath.stem + "_step0002")
    return objBasePath.with_suffix(".xlsx"), objBasePath.with_suffix(".tsv")


def get_step0003_output_file_paths(pszInputFileFullPath: str) -> tuple[Path, Path]:
    """_step0003.xlsxと_step0003.tsvの出力パスを返します。"""
    objInputPath: Path = Path(pszInputFileFullPath)
    objBasePath: Path = objInputPath.with_name(objInputPath.stem + "_step0003")
    return objBasePath.with_suffix(".xlsx"), objBasePath.with_suffix(".tsv")


def get_step0004_output_file_paths(pszInputFileFullPath: str) -> tuple[Path, Path]:
    """_step0004.xlsxと_step0004.tsvの出力パスを返します。"""
    objInputPath: Path = Path(pszInputFileFullPath)
    objBasePath: Path = objInputPath.with_name(objInputPath.stem + "_step0004")
    return objBasePath.with_suffix(".xlsx"), objBasePath.with_suffix(".tsv")


def create_temporary_path(objOutputPath: Path, pszSuffix: str) -> Path:
    """出力先と同じフォルダーに一意な一時ファイルパスを作ります。"""
    iFileDescriptor, pszTemporaryPath = tempfile.mkstemp(
        prefix=objOutputPath.stem + "_", suffix=pszSuffix, dir=objOutputPath.parent
    )
    os.close(iFileDescriptor)
    return Path(pszTemporaryPath)


def save_excel_template(objOutputPath: Path, listProductRows: list[ProductRow]) -> None:
    """3列の新規Excelテンプレートを保存します。"""
    objWorkbook: Workbook = Workbook()
    objWorksheet: Worksheet = objWorkbook.active
    objWorksheet.append(list(EXPECTED_HEADERS))
    for objProductRow in listProductRows:
        objWorksheet.append(
            [objProductRow.product_code, objProductRow.product_name, objProductRow.spec]
        )
        objWorksheet.cell(row=objWorksheet.max_row, column=1).number_format = "@"
    objWorkbook.save(objOutputPath)


def save_tsv_template(objOutputPath: Path, listProductRows: list[ProductRow]) -> None:
    """3列のTSVをUTF-8 BOMなし、CRLFで保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow(EXPECTED_HEADERS)
        for objProductRow in listProductRows:
            objWriter.writerow(
                [objProductRow.product_code, objProductRow.product_name, objProductRow.spec]
            )


def build_step0002_row(objProductRow: ProductRow) -> list[str]:
    """処理0001の商品行を処理0002の14列へ変換します。"""
    return [
        "",
        "",
        "",
        "",
        objProductRow.product_code,
        objProductRow.product_code,
        objProductRow.product_name,
        "",
        objProductRow.spec,
        "",
        "",
        "",
        "",
        "",
    ]


def save_step0002_excel_template(
    objOutputPath: Path, listProductRows: list[ProductRow]
) -> None:
    """処理0002の14列を持つ新規Excelテンプレートを保存します。"""
    objWorkbook: Workbook = Workbook()
    objWorksheet: Worksheet = objWorkbook.active
    objWorksheet.append(list(STEP0002_HEADERS))
    for objProductRow in listProductRows:
        objWorksheet.append(build_step0002_row(objProductRow))
        iOutputRow: int = objWorksheet.max_row
        objWorksheet.cell(row=iOutputRow, column=5).number_format = "@"
        objWorksheet.cell(row=iOutputRow, column=6).number_format = "@"
    objWorkbook.save(objOutputPath)


def save_step0002_tsv_template(
    objOutputPath: Path, listProductRows: list[ProductRow]
) -> None:
    """処理0002の14列TSVをUTF-8 BOMなし、CRLFで保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow(STEP0002_HEADERS)
        for objProductRow in listProductRows:
            objWriter.writerow(build_step0002_row(objProductRow))


def validate_step0001_outputs_match(
    listExcelRows: list[ProductRow], listTsvRows: list[ProductRow]
) -> None:
    """処理0001のXLSXとTSVの行数・行順・各値が一致することを確認します。"""
    if len(listExcelRows) != len(listTsvRows):
        raise ValueError(
            "step0001のXLSXとTSVの内容が一致しません。データ行数: XLSX = "
            + str(len(listExcelRows))
            + "、TSV = "
            + str(len(listTsvRows))
        )
    tupleColumns: tuple[tuple[str, str], ...] = (
        ("productCode", "product_code"),
        ("productName", "product_name"),
        ("spec", "spec"),
    )
    for iRow, (objExcelRow, objTsvRow) in enumerate(
        zip(listExcelRows, listTsvRows), start=2
    ):
        for pszColumnName, pszAttributeName in tupleColumns:
            pszExcelValue: str = getattr(objExcelRow, pszAttributeName)
            pszTsvValue: str = getattr(objTsvRow, pszAttributeName)
            if pszExcelValue != pszTsvValue:
                raise ValueError(
                    "step0001のXLSXとTSVの内容が一致しません。行 = "
                    + str(iRow)
                    + "、列 = "
                    + pszColumnName
                    + "、XLSX = "
                    + repr(pszExcelValue)
                    + "、TSV = "
                    + repr(pszTsvValue)
                )


def normalize_template_row(listValues: list[object]) -> list[str]:
    """処理0002の14列を、空セルを空文字にした比較・出力用文字列へ変換します。"""
    return [normalize_text(objValue) for objValue in listValues[: len(STEP0002_HEADERS)]]


def read_step0002_excel_rows(
    objExcelPath: Path, pszStepName: str = "step0002"
) -> list[list[str]]:
    """14列テンプレートExcelを検証してデータ行を読み取ります。"""
    objWorkbook: Workbook = load_workbook(objExcelPath, data_only=True)
    listTargetWorksheets: list[Worksheet] = []
    for objWorksheet in objWorkbook.worksheets:
        tupleHeaders: tuple[str, ...] = tuple(
            normalize_header(objWorksheet.cell(row=1, column=iColumn).value)
            for iColumn in range(1, objWorksheet.max_column + 1)
        )
        if tupleHeaders == STEP0002_HEADERS:
            listTargetWorksheets.append(objWorksheet)
    if len(listTargetWorksheets) == 0:
        raise ValueError(
            pszStepName + "の14列ヘッダーを持つExcelシートが見つかりません。"
        )
    if len(listTargetWorksheets) > 1:
        raise ValueError(
            pszStepName
            + "の対象シートが複数見つかりました。対象シート = "
            + ", ".join(objWorksheet.title for objWorksheet in listTargetWorksheets)
        )
    objWorksheet: Worksheet = listTargetWorksheets[0]
    listRows: list[list[str]] = []
    for iRow in range(2, objWorksheet.max_row + 1):
        listValues: list[object] = [
            objWorksheet.cell(row=iRow, column=iColumn).value
            for iColumn in range(1, len(STEP0002_HEADERS) + 1)
        ]
        listNormalizedRow: list[str] = normalize_template_row(listValues)
        if all(pszValue.strip() == "" for pszValue in listNormalizedRow):
            continue
        listRows.append(listNormalizedRow)
    return listRows


def read_step0002_tsv_rows(
    objTsvPath: Path, pszStepName: str = "step0002"
) -> list[list[str]]:
    """14列テンプレートのUTF-8 TSVを検証してデータ行を読み取ります。"""
    with objTsvPath.open(mode="r", encoding="utf-8", newline="") as objFile:
        listRows: list[list[str]] = list(
            csv.reader(objFile, delimiter="\t", strict=True)
        )
    if not listRows:
        raise ValueError(pszStepName + "のTSVファイルが空です。")
    if tuple(normalize_header(pszValue) for pszValue in listRows[0]) != STEP0002_HEADERS:
        raise ValueError(
            pszStepName + "のTSVヘッダーが仕様どおりの14列ではありません。"
        )
    listNormalizedRows: list[list[str]] = []
    for iRow, listValues in enumerate(listRows[1:], start=2):
        if len(listValues) != len(STEP0002_HEADERS):
            raise ValueError(
                str(iRow)
                + "行目の列数が14列ではありません。列数 = "
                + str(len(listValues))
            )
        listNormalizedRow: list[str] = normalize_template_row(listValues)
        if all(pszValue.strip() == "" for pszValue in listNormalizedRow):
            continue
        listNormalizedRows.append(listNormalizedRow)
    return listNormalizedRows


def validate_step0002_outputs_match(
    listExcelRows: list[list[str]],
    listTsvRows: list[list[str]],
    pszStepName: str = "step0002",
) -> None:
    """14列XLSX・TSVの行数、行順、すべての値が一致するか確認します。"""
    if len(listExcelRows) != len(listTsvRows):
        raise ValueError(
            pszStepName
            + "のXLSXとTSVの内容が一致しません。データ行数: XLSX = "
            + str(len(listExcelRows))
            + "、TSV = "
            + str(len(listTsvRows))
        )
    for iRow, (listExcelRow, listTsvRow) in enumerate(
        zip(listExcelRows, listTsvRows), start=2
    ):
        for iColumn, pszColumnName in enumerate(STEP0002_HEADERS):
            if listExcelRow[iColumn] != listTsvRow[iColumn]:
                raise ValueError(
                    pszStepName
                    + "のXLSXとTSVの内容が一致しません。行 = "
                    + str(iRow)
                    + "、列 = "
                    + pszColumnName
                    + "、XLSX = "
                    + repr(listExcelRow[iColumn])
                    + "、TSV = "
                    + repr(listTsvRow[iColumn])
                )


def build_step0003_rows(listStep0002Rows: list[list[str]]) -> list[list[str]]:
    """処理0002の商品1行を月～日の7行へ展開します。"""
    listStep0003Rows: list[list[str]] = []
    for listStep0002Row in listStep0002Rows:
        listMondayRow: list[str] = listStep0002Row.copy()
        listMondayRow[1] = WEEKDAYS[0]
        listStep0003Rows.append(listMondayRow)
        for pszWeekday in WEEKDAYS[1:]:
            listWeekdayRow: list[str] = [""] * len(STEP0002_HEADERS)
            listWeekdayRow[1] = pszWeekday
            listStep0003Rows.append(listWeekdayRow)
    return listStep0003Rows


def save_step0003_excel_template(
    objOutputPath: Path, listStep0003Rows: list[list[str]]
) -> None:
    """月～日に展開した処理0003のExcelテンプレートを保存します。"""
    objWorkbook: Workbook = Workbook()
    objWorksheet: Worksheet = objWorkbook.active
    objWorksheet.append(list(STEP0002_HEADERS))
    for listValues in listStep0003Rows:
        objWorksheet.append(listValues)
        if listValues[1] == WEEKDAYS[0]:
            iOutputRow: int = objWorksheet.max_row
            objWorksheet.cell(row=iOutputRow, column=5).number_format = "@"
            objWorksheet.cell(row=iOutputRow, column=6).number_format = "@"
    objWorkbook.save(objOutputPath)


def save_step0003_tsv_template(
    objOutputPath: Path, listStep0003Rows: list[list[str]]
) -> None:
    """月～日に展開した処理0003のTSVをUTF-8 BOMなし、CRLFで保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow(STEP0002_HEADERS)
        objWriter.writerows(listStep0003Rows)


def validate_step0003_weekday_cycle(listStep0003Rows: list[list[str]]) -> None:
    """処理0003が7行周期で月～日の順になっていることを確認します。"""
    if len(listStep0003Rows) % len(WEEKDAYS) != 0:
        raise ValueError(
            "step0003のデータ行数が7の倍数ではありません。データ行数 = "
            + str(len(listStep0003Rows))
        )
    for iRowIndex, listValues in enumerate(listStep0003Rows):
        pszExpectedWeekday: str = WEEKDAYS[iRowIndex % len(WEEKDAYS)]
        pszActualWeekday: str = listValues[1]
        if pszActualWeekday != pszExpectedWeekday:
            raise ValueError(
                "step0003の曜日順が仕様と一致しません。行 = "
                + str(iRowIndex + 2)
                + "、期待値 = "
                + pszExpectedWeekday
                + "、実際値 = "
                + pszActualWeekday
            )


def build_step0004_rows(
    listStep0003Rows: list[list[str]], objStartMonday: date
) -> list[list[str]]:
    """7行ごとに同じ月～日の年月日を納品日列へ設定します。"""
    validate_start_monday(objStartMonday)
    validate_step0003_weekday_cycle(listStep0003Rows)
    listStep0004Rows: list[list[str]] = []
    for iRowIndex, listStep0003Row in enumerate(listStep0003Rows):
        objDeliveryDate: date = objStartMonday + timedelta(
            days=iRowIndex % len(WEEKDAYS)
        )
        listStep0004Row: list[str] = listStep0003Row.copy()
        listStep0004Row[0] = objDeliveryDate.strftime("%Y/%m/%d")
        listStep0004Rows.append(listStep0004Row)
    return listStep0004Rows


def save_step0004_excel_template(
    objOutputPath: Path, listStep0004Rows: list[list[str]]
) -> None:
    """納品日をExcel日付値として持つ処理0004テンプレートを保存します。"""
    objWorkbook: Workbook = Workbook()
    objWorksheet: Worksheet = objWorkbook.active
    objWorksheet.append(list(STEP0002_HEADERS))
    for listValues in listStep0004Rows:
        listExcelValues: list[object] = listValues.copy()
        listExcelValues[0] = datetime.strptime(listValues[0], "%Y/%m/%d").date()
        objWorksheet.append(listExcelValues)
        iOutputRow: int = objWorksheet.max_row
        objWorksheet.cell(row=iOutputRow, column=1).number_format = "yyyy/mm/dd"
        if listValues[1] == WEEKDAYS[0]:
            objWorksheet.cell(row=iOutputRow, column=5).number_format = "@"
            objWorksheet.cell(row=iOutputRow, column=6).number_format = "@"
    objWorkbook.save(objOutputPath)


def save_step0004_tsv_template(
    objOutputPath: Path, listStep0004Rows: list[list[str]]
) -> None:
    """処理0004のTSVをUTF-8 BOMなし、タブ区切り、CRLFで保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow(STEP0002_HEADERS)
        objWriter.writerows(listStep0004Rows)


def replace_output_files(
    objTemporaryExcelPath: Path,
    objTemporaryTsvPath: Path,
    objExcelOutputPath: Path,
    objTsvOutputPath: Path,
) -> None:
    """2出力を置換し、失敗時は可能な限り以前の状態へ戻します。"""
    listOutputPaths: list[Path] = [objExcelOutputPath, objTsvOutputPath]
    listTemporaryPaths: list[Path] = [objTemporaryExcelPath, objTemporaryTsvPath]
    dictBackupPaths: dict[Path, Path] = {}
    listReplacedPaths: list[Path] = []
    try:
        for objOutputPath in listOutputPaths:
            if objOutputPath.exists():
                objBackupPath: Path = create_temporary_path(objOutputPath, ".backup")
                shutil.copy2(objOutputPath, objBackupPath)
                dictBackupPaths[objOutputPath] = objBackupPath
        for objTemporaryPath, objOutputPath in zip(listTemporaryPaths, listOutputPaths):
            os.replace(objTemporaryPath, objOutputPath)
            listReplacedPaths.append(objOutputPath)
    except Exception:
        for objOutputPath in reversed(listReplacedPaths):
            objBackupPath = dictBackupPaths.get(objOutputPath)
            if objBackupPath is not None and objBackupPath.exists():
                os.replace(objBackupPath, objOutputPath)
            elif objOutputPath.exists():
                objOutputPath.unlink()
        raise
    finally:
        for objBackupPath in dictBackupPaths.values():
            if objBackupPath.exists():
                objBackupPath.unlink()
        for objTemporaryPath in listTemporaryPaths:
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()


def process_step0002_files(
    pszInputFileFullPath: str,
    objStep0001ExcelPath: Path,
    objStep0001TsvPath: Path,
) -> tuple[Path, Path, int]:
    """処理0001の両出力を比較し、処理0002のXLSXとTSVを作成します。"""
    try:
        listExcelRows: list[ProductRow] = read_excel_rows(str(objStep0001ExcelPath))
        listTsvRows: list[ProductRow] = read_delimited_rows(str(objStep0001TsvPath))
        validate_step0001_outputs_match(listExcelRows, listTsvRows)

        objExcelOutputPath, objTsvOutputPath = get_step0002_output_file_paths(
            pszInputFileFullPath
        )
        objTemporaryExcelPath: Path = create_temporary_path(objExcelOutputPath, ".xlsx")
        objTemporaryTsvPath: Path = create_temporary_path(objTsvOutputPath, ".tsv")
        try:
            save_step0002_excel_template(objTemporaryExcelPath, listExcelRows)
            save_step0002_tsv_template(objTemporaryTsvPath, listExcelRows)
            replace_output_files(
                objTemporaryExcelPath,
                objTemporaryTsvPath,
                objExcelOutputPath,
                objTsvOutputPath,
            )
        finally:
            for objTemporaryPath in (objTemporaryExcelPath, objTemporaryTsvPath):
                if objTemporaryPath.exists():
                    objTemporaryPath.unlink()
        return objExcelOutputPath, objTsvOutputPath, len(listExcelRows)
    except Exception as objException:
        raise Step0002Error(str(objException)) from objException


def process_step0003_files(
    pszInputFileFullPath: str,
    objStep0002ExcelPath: Path,
    objStep0002TsvPath: Path,
) -> tuple[Path, Path, int, int]:
    """処理0002の両出力を比較し、月～日に展開した処理0003を作成します。"""
    try:
        listExcelRows: list[list[str]] = read_step0002_excel_rows(
            objStep0002ExcelPath
        )
        listTsvRows: list[list[str]] = read_step0002_tsv_rows(objStep0002TsvPath)
        validate_step0002_outputs_match(listExcelRows, listTsvRows)
        listStep0003Rows: list[list[str]] = build_step0003_rows(listExcelRows)

        objExcelOutputPath, objTsvOutputPath = get_step0003_output_file_paths(
            pszInputFileFullPath
        )
        objTemporaryExcelPath: Path = create_temporary_path(objExcelOutputPath, ".xlsx")
        objTemporaryTsvPath: Path = create_temporary_path(objTsvOutputPath, ".tsv")
        try:
            save_step0003_excel_template(objTemporaryExcelPath, listStep0003Rows)
            save_step0003_tsv_template(objTemporaryTsvPath, listStep0003Rows)
            replace_output_files(
                objTemporaryExcelPath,
                objTemporaryTsvPath,
                objExcelOutputPath,
                objTsvOutputPath,
            )
        finally:
            for objTemporaryPath in (objTemporaryExcelPath, objTemporaryTsvPath):
                if objTemporaryPath.exists():
                    objTemporaryPath.unlink()
        return (
            objExcelOutputPath,
            objTsvOutputPath,
            len(listExcelRows),
            len(listStep0003Rows),
        )
    except Exception as objException:
        raise Step0003Error(str(objException)) from objException


def process_step0004_files(
    pszInputFileFullPath: str,
    objStep0003ExcelPath: Path,
    objStep0003TsvPath: Path,
    objStartMonday: date,
) -> tuple[Path, Path, int]:
    """処理0003の両出力を比較し、同じ1週間を繰り返す処理0004を作成します。"""
    try:
        listExcelRows: list[list[str]] = read_step0002_excel_rows(
            objStep0003ExcelPath, "step0003"
        )
        listTsvRows: list[list[str]] = read_step0002_tsv_rows(
            objStep0003TsvPath, "step0003"
        )
        validate_step0002_outputs_match(listExcelRows, listTsvRows, "step0003")
        listStep0004Rows: list[list[str]] = build_step0004_rows(
            listExcelRows, objStartMonday
        )

        objExcelOutputPath, objTsvOutputPath = get_step0004_output_file_paths(
            pszInputFileFullPath
        )
        objTemporaryExcelPath: Path = create_temporary_path(objExcelOutputPath, ".xlsx")
        objTemporaryTsvPath: Path = create_temporary_path(objTsvOutputPath, ".tsv")
        try:
            save_step0004_excel_template(objTemporaryExcelPath, listStep0004Rows)
            save_step0004_tsv_template(objTemporaryTsvPath, listStep0004Rows)
            replace_output_files(
                objTemporaryExcelPath,
                objTemporaryTsvPath,
                objExcelOutputPath,
                objTsvOutputPath,
            )
        finally:
            for objTemporaryPath in (objTemporaryExcelPath, objTemporaryTsvPath):
                if objTemporaryPath.exists():
                    objTemporaryPath.unlink()
        return objExcelOutputPath, objTsvOutputPath, len(listStep0004Rows)
    except Exception as objException:
        raise Step0004Error(str(objException)) from objException


def process_input_file(pszInputFileFullPath: str, objStartMonday: date) -> None:
    """入力から処理0001～処理0004のXLSX・TSVを作成します。"""
    validate_start_monday(objStartMonday)
    pszValidatedPath: str = validate_input_path(pszInputFileFullPath)
    pszExtension: str = os.path.splitext(pszValidatedPath)[1].lower()
    if pszExtension == ".xlsx":
        listProductRows: list[ProductRow] = read_excel_rows(pszValidatedPath)
    else:
        listProductRows = read_delimited_rows(pszValidatedPath)
    objExcelOutputPath, objTsvOutputPath = get_output_file_paths(pszValidatedPath)
    objTemporaryExcelPath: Path = create_temporary_path(objExcelOutputPath, ".xlsx")
    objTemporaryTsvPath: Path = create_temporary_path(objTsvOutputPath, ".tsv")
    try:
        save_excel_template(objTemporaryExcelPath, listProductRows)
        save_tsv_template(objTemporaryTsvPath, listProductRows)
        replace_output_files(
            objTemporaryExcelPath,
            objTemporaryTsvPath,
            objExcelOutputPath,
            objTsvOutputPath,
        )
    finally:
        for objTemporaryPath in (objTemporaryExcelPath, objTemporaryTsvPath):
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()

    objStep0002ExcelPath, objStep0002TsvPath, _ = (
        process_step0002_files(
            pszValidatedPath,
            objExcelOutputPath,
            objTsvOutputPath,
        )
    )
    (
        objStep0003ExcelPath,
        objStep0003TsvPath,
        iProductCount,
        _,
    ) = process_step0003_files(
        pszValidatedPath,
        objStep0002ExcelPath,
        objStep0002TsvPath,
    )
    objStep0004ExcelPath, objStep0004TsvPath, iStep0004RowCount = (
        process_step0004_files(
            pszValidatedPath,
            objStep0003ExcelPath,
            objStep0003TsvPath,
            objStartMonday,
        )
    )
    remove_old_error_file(pszValidatedPath)
    print("朝日注文テンプレートファイルを作成しました。")
    print("Input: " + pszValidatedPath)
    print("Start Monday: " + objStartMonday.strftime("%Y/%m/%d"))
    print("Step0001 Excel: " + str(objExcelOutputPath))
    print("Step0001 TSV: " + str(objTsvOutputPath))
    print("Step0002 Excel: " + str(objStep0002ExcelPath))
    print("Step0002 TSV: " + str(objStep0002TsvPath))
    print("Step0003 Excel: " + str(objStep0003ExcelPath))
    print("Step0003 TSV: " + str(objStep0003TsvPath))
    print("Step0004 Excel: " + str(objStep0004ExcelPath))
    print("Step0004 TSV: " + str(objStep0004TsvPath))
    print("Products: " + str(iProductCount))
    print("Step0004 Rows: " + str(iStep0004RowCount))


def main() -> int:
    """引数を確認して処理し、成功0・失敗1の終了コードを返します。"""
    bHasStartMondayArgument: bool = len(sys.argv) == 4 and sys.argv[1] == "--start-monday"
    bHasOnlyInputArgument: bool = len(sys.argv) == 2
    if not bHasStartMondayArgument and not bHasOnlyInputArgument:
        pszScriptFileName: str = os.path.basename(__file__)
        pszErrorMessage: str = (
            "Error: 入力ファイルパスと開始月曜日を正しく指定してください。\n"
            + "Usage: python "
            + pszScriptFileName
            + " <input_file_path>\n"
            + "With date: python "
            + pszScriptFileName
            + " --start-monday YYYY-MM-DD <input_file_path>\n"
        )
        print(pszErrorMessage, file=sys.stderr, end="")
        pszErrorFileFullPath: str = os.path.splitext(pszScriptFileName)[0] + "_error_argument.txt"
        try:
            write_error_text(pszErrorFileFullPath, pszErrorMessage)
        except OSError as objException:
            print(
                "Error: 引数エラーファイルを保存できません。Detail = " + str(objException),
                file=sys.stderr,
            )
        return 1
    pszInputFileFullPath: str = sys.argv[3] if bHasStartMondayArgument else sys.argv[1]
    try:
        try:
            if bHasStartMondayArgument:
                objStartMonday: date = parse_start_monday(sys.argv[2])
            else:
                objSelectedMonday: date | None = select_start_monday()
                if objSelectedMonday is None:
                    show_start_monday_cancelled_message()
                    raise Step0004Error(
                        "開始月曜日の選択がキャンセルされたため、処理0004を中止しました。"
                    )
                objStartMonday = objSelectedMonday
        except Step0004Error:
            raise
        except Exception as objException:
            raise Step0004Error(str(objException)) from objException
        process_input_file(pszInputFileFullPath, objStartMonday)
    except Step0004Error as objException:
        report_processing_error(
            pszInputFileFullPath,
            "朝日注文テンプレート処理0004",
            str(objException),
        )
        return 1
    except Step0003Error as objException:
        report_processing_error(
            pszInputFileFullPath,
            "朝日注文テンプレート処理0003",
            str(objException),
        )
        return 1
    except Step0002Error as objException:
        report_processing_error(
            pszInputFileFullPath,
            "朝日注文テンプレート処理0002",
            str(objException),
        )
        return 1
    except Exception as objException:
        report_processing_error(
            pszInputFileFullPath,
            "朝日注文テンプレート処理0001",
            str(objException),
        )
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
