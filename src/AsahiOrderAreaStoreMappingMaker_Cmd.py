# -- coding: utf-8 --
###############################################################
#
# AsahiOrderAreaStoreMappingMaker_Cmd.py
#
# pip install openpyxl
#
###############################################################

import csv
import os
import re
import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


TARGET_WORKSHEET_NAMES: tuple[str, str] = ("本州マグロ(週間)", "割り")
SUPPORTED_EXTENSIONS: set[str] = {".xlsx"}
MAX_BACKUP_NUMBER: int = 9999
AREA_STORE_MAPPING_FILE_NAME: str = "AsahiOrderAreaStoreMapping_週間_step0001.tsv"
AREA_STORE_MAPPING_HEADERS: tuple[str, str, str] = (
    "配送センター名",
    "店舗コード",
    "店舗名",
)
ALLOCATION_MAPPING_FILE_NAME: str = "AsahiOrderAreaStoreMapping_割り_step0001.tsv"
ALLOCATION_MAPPING_HEADERS: tuple[str, str, str, str, str] = (
    "エリア名",
    "店舗コード",
    "店舗略称",
    "APEX店舗コード",
    "APEX店舗名",
)
CIRCLED_NUMBERS: str = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳"


def write_error_text(pszOutputFileFullPath: str, pszErrorMessage: str) -> None:
    """エラーメッセージをUTF-8のテキストファイルへ上書き保存します。"""
    pszDirectoryFullPath: str = os.path.dirname(pszOutputFileFullPath)
    if pszDirectoryFullPath != "":
        os.makedirs(pszDirectoryFullPath, exist_ok=True)
    with open(pszOutputFileFullPath, mode="w", encoding="utf-8", newline="") as objFile:
        objFile.write(pszErrorMessage.rstrip("\n") + "\n")


def get_error_file_full_path(pszInputFileFullPath: str) -> str:
    """入力ファイルと同じフォルダーに作る_error.txtのパスを返します。"""
    pszDirectoryFullPath: str = os.path.dirname(os.path.abspath(pszInputFileFullPath))
    pszBaseNameWithoutExtension: str = os.path.splitext(
        os.path.basename(pszInputFileFullPath)
    )[0]
    return os.path.join(pszDirectoryFullPath, pszBaseNameWithoutExtension + "_error.txt")


def report_processing_error(
    pszInputFileFullPath: str, pszProcessName: str, pszDetailMessage: str
) -> None:
    """標準エラーと入力ファイル用_error.txtへ同じエラーを出力します。"""
    pszErrorMessage: str = (
        "処理結果: エラー\n"
        + "入力ファイル: "
        + os.path.abspath(pszInputFileFullPath)
        + "\n発生した処理: "
        + pszProcessName
        + "\nエラー内容: "
        + pszDetailMessage
        + "\n"
    )
    print(pszErrorMessage, file=sys.stderr, end="")
    try:
        write_error_text(get_error_file_full_path(pszInputFileFullPath), pszErrorMessage)
    except OSError as objException:
        print(
            "Error: _error.txtの保存にも失敗しました。Detail = " + str(objException),
            file=sys.stderr,
        )
    try:
        pszWarningFileFullPath: str = get_warning_file_full_path(
            pszInputFileFullPath
        )
        if os.path.exists(pszWarningFileFullPath):
            os.remove(pszWarningFileFullPath)
    except OSError as objException:
        print(
            "Error: 古い_warning.txtを削除できませんでした。Detail = "
            + str(objException),
            file=sys.stderr,
        )


def remove_old_error_file(pszInputFileFullPath: str) -> None:
    """正常終了後、以前の処理で作られた_error.txtがあれば削除します。"""
    pszErrorFileFullPath: str = get_error_file_full_path(pszInputFileFullPath)
    if os.path.exists(pszErrorFileFullPath):
        os.remove(pszErrorFileFullPath)


def get_warning_file_full_path(pszInputFileFullPath: str) -> str:
    """入力ファイルと同じフォルダーに作る_warning.txtのパスを返します。"""
    pszDirectoryFullPath: str = os.path.dirname(os.path.abspath(pszInputFileFullPath))
    pszBaseNameWithoutExtension: str = os.path.splitext(
        os.path.basename(pszInputFileFullPath)
    )[0]
    return os.path.join(
        pszDirectoryFullPath, pszBaseNameWithoutExtension + "_warning.txt"
    )


def validate_input_path(pszInputFileFullPath: str) -> str:
    """入力パスと拡張子を検証し、絶対パスを返します。"""
    pszAbsolutePath: str = os.path.abspath(pszInputFileFullPath)
    if not os.path.exists(pszAbsolutePath):
        raise ValueError("入力ファイルが見つかりません。Path = " + pszAbsolutePath)
    if not os.path.isfile(pszAbsolutePath):
        raise ValueError("入力パスがファイルではありません。Path = " + pszAbsolutePath)
    pszExtension: str = os.path.splitext(pszAbsolutePath)[1].lower()
    if pszExtension not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            "入力ファイルの拡張子は.xlsxではありません。Path = " + pszAbsolutePath
        )
    return pszAbsolutePath


def normalize_cell_value(objValue: object) -> object:
    """空セルを空文字へ変換し、それ以外のセル値は変更せず返します。"""
    if objValue is None:
        return ""
    return objValue


def is_empty_cell_value(objValue: object) -> bool:
    """末尾空行・空列の判定対象となる空セルかを返します。"""
    return objValue is None or objValue == ""


def read_worksheet_rows(objWorksheet: Worksheet) -> tuple[list[list[object]], int]:
    """シートのセル値を読み、末尾の完全空行・空列を除いて返します。"""
    iLastDataRow: int = 0
    iLastDataColumn: int = 0
    for objRow in objWorksheet.iter_rows(
        min_row=1,
        max_row=objWorksheet.max_row,
        min_col=1,
        max_col=objWorksheet.max_column,
    ):
        for objCell in objRow:
            if not is_empty_cell_value(objCell.value):
                iLastDataRow = max(iLastDataRow, objCell.row)
                iLastDataColumn = max(iLastDataColumn, objCell.column)
    if iLastDataRow == 0 or iLastDataColumn == 0:
        return [], 0
    listRows: list[list[object]] = []
    for objRow in objWorksheet.iter_rows(
        min_row=1,
        max_row=iLastDataRow,
        min_col=1,
        max_col=iLastDataColumn,
    ):
        listRows.append([normalize_cell_value(objCell.value) for objCell in objRow])
    return listRows, iLastDataColumn


def read_excel_worksheets(
    pszInputFileFullPath: str,
) -> tuple[dict[str, tuple[list[list[object]], int]], list[str]]:
    """対象シートを1回のExcel読込で取得し、存在しないシート名も返します。"""
    objWorkbook: Workbook = load_workbook(pszInputFileFullPath, data_only=True)
    try:
        dictWorksheetResults: dict[str, tuple[list[list[object]], int]] = {}
        listMissingWorksheetNames: list[str] = []
        for pszWorksheetName in TARGET_WORKSHEET_NAMES:
            if pszWorksheetName not in objWorkbook.sheetnames:
                listMissingWorksheetNames.append(pszWorksheetName)
                continue
            dictWorksheetResults[pszWorksheetName] = read_worksheet_rows(
                objWorkbook[pszWorksheetName]
            )
        return dictWorksheetResults, listMissingWorksheetNames
    finally:
        objWorkbook.close()


def get_output_file_path(
    pszInputFileFullPath: str, pszWorksheetName: str
) -> Path:
    """入力ファイル名とシート名からTSVの出力パスを返します。"""
    objInputPath: Path = Path(pszInputFileFullPath)
    return objInputPath.with_name(objInputPath.stem + "_" + pszWorksheetName + ".tsv")


def create_temporary_path(objOutputPath: Path) -> Path:
    """出力先と同じフォルダーに同じ拡張子の一時パスを作ります。"""
    iFileDescriptor, pszTemporaryPath = tempfile.mkstemp(
        prefix=objOutputPath.stem + "_",
        suffix=objOutputPath.suffix,
        dir=objOutputPath.parent,
    )
    os.close(iFileDescriptor)
    return Path(pszTemporaryPath)


def save_tsv_rows(objOutputPath: Path, listRows: list[list[object]]) -> None:
    """セル値をUTF-8 BOMなし、タブ区切り、CRLFのTSVへ保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerows(listRows)


def read_tsv_rows(objInputPath: Path) -> list[list[str]]:
    """UTF-8のTSVを先頭から終端まで読み取ります。"""
    with objInputPath.open(mode="r", encoding="utf-8", newline="") as objFile:
        return list(csv.reader(objFile, delimiter="\t", strict=True))


def is_allocation_store_data_row(listValues: list[str]) -> bool:
    """割り対応表の2～6列目がすべて空白以外の店舗行か返します。"""
    if len(listValues) < 6:
        return False
    return all(pszValue.strip() != "" for pszValue in listValues[1:6])


def build_allocation_mapping_rows(listRows: list[list[str]]) -> list[list[str]]:
    """単位行より後ろから割り対応表の2～6列目を抽出します。"""
    iUnitRowIndex: int | None = None
    for iRowIndex, listValues in enumerate(listRows):
        if len(listValues) >= 6 and listValues[5] == "単位":
            iUnitRowIndex = iRowIndex
            break
    if iUnitRowIndex is None:
        raise ValueError("割りTSV内に6列目が「単位」の行が見つかりません。")
    iDataStartRowIndex: int | None = None
    for iRowIndex in range(iUnitRowIndex + 1, len(listRows)):
        if is_allocation_store_data_row(listRows[iRowIndex]):
            iDataStartRowIndex = iRowIndex
            break
    if iDataStartRowIndex is None:
        raise ValueError(
            "6列目が「単位」の行より後ろに店舗データが見つかりません。"
        )
    return [
        listValues[1:6]
        for listValues in listRows[iDataStartRowIndex:]
        if is_allocation_store_data_row(listValues)
    ]


def get_allocation_mapping_output_path(objInputTsvPath: Path) -> Path:
    """割りTSVと同じフォルダーに作る割り対応表TSVのパスを返します。"""
    return objInputTsvPath.with_name(ALLOCATION_MAPPING_FILE_NAME)


def save_allocation_mapping_tsv(
    objOutputPath: Path, listMappingRows: list[list[str]]
) -> None:
    """5列ヘッダーと割り対応データをUTF-8 TSVへ保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow(ALLOCATION_MAPPING_HEADERS)
        objWriter.writerows(listMappingRows)


def process_allocation_mapping_file(objInputTsvPath: Path) -> tuple[Path, int]:
    """割りTSVから割り対応表を独立して作成します。"""
    if not objInputTsvPath.exists() or not objInputTsvPath.is_file():
        raise ValueError("割りTSVが見つかりません。Path = " + str(objInputTsvPath))
    listRows: list[list[str]] = read_tsv_rows(objInputTsvPath)
    try:
        listMappingRows: list[list[str]] = build_allocation_mapping_rows(listRows)
    except ValueError as objException:
        raise ValueError(
            str(objException) + " Path = " + str(objInputTsvPath)
        ) from objException
    objOutputPath: Path = get_allocation_mapping_output_path(objInputTsvPath)
    objTemporaryPath: Path = create_temporary_path(objOutputPath)
    try:
        save_allocation_mapping_tsv(objTemporaryPath, listMappingRows)
        os.replace(objTemporaryPath, objOutputPath)
    finally:
        if objTemporaryPath.exists():
            objTemporaryPath.unlink()
    return objOutputPath, len(listMappingRows)


def get_row_value(listValues: list[str], iColumnIndex: int) -> str:
    """指定列が存在すれば値を返し、列不足なら空文字を返します。"""
    if iColumnIndex >= len(listValues):
        return ""
    return listValues[iColumnIndex]


def normalize_distribution_center_name(pszValue: str) -> str:
    """配送センター名の前後空白を除き、括弧を半角へ統一します。"""
    return pszValue.strip().replace("（", "(").replace("）", ")")


def is_distribution_center_name(pszValue: str) -> bool:
    """丸数字で始まりセンターを含む配送センター見出しか返します。"""
    pszNormalizedValue: str = normalize_distribution_center_name(pszValue)
    return (
        pszNormalizedValue != ""
        and pszNormalizedValue[0] in CIRCLED_NUMBERS
        and "センター" in pszNormalizedValue
    )


def find_distribution_centers(
    listRows: list[list[str]],
) -> list[tuple[int, int, str]]:
    """全セルから配送センター見出しの行・列・正規化名を返します。"""
    listCenters: list[tuple[int, int, str]] = []
    for iRowIndex, listValues in enumerate(listRows):
        for iColumnIndex, pszValue in enumerate(listValues):
            if is_distribution_center_name(pszValue):
                listCenters.append(
                    (
                        iRowIndex,
                        iColumnIndex,
                        normalize_distribution_center_name(pszValue),
                    )
                )
    if not listCenters:
        raise ValueError("配送センター見出しが見つかりません。")
    return listCenters


def find_store_groups(listRows: list[list[str]]) -> list[tuple[int, int, int]]:
    """店舗コードと末尾が店舗名の隣接ヘッダーをすべて返します。"""
    listStoreGroups: list[tuple[int, int, int]] = []
    for iRowIndex, listValues in enumerate(listRows):
        for iColumnIndex, pszValue in enumerate(listValues):
            if pszValue != "店舗コード":
                continue
            pszStoreNameHeader: str = get_row_value(listValues, iColumnIndex + 1)
            if pszStoreNameHeader.endswith("店舗名"):
                listStoreGroups.append(
                    (iRowIndex, iColumnIndex, iColumnIndex + 1)
                )
    if not listStoreGroups:
        raise ValueError("店舗コード・店舗名の組が見つかりません。")
    return listStoreGroups


def assign_distribution_center(
    tupleStoreGroup: tuple[int, int, int],
    listCenters: list[tuple[int, int, str]],
) -> tuple[int, int, str]:
    """店舗グループの上方で最も近い見出し行から所属センターを返します。"""
    iHeaderRowIndex, iStoreCodeColumnIndex, _ = tupleStoreGroup
    listPreviousCenters: list[tuple[int, int, str]] = [
        tupleCenter
        for tupleCenter in listCenters
        if tupleCenter[0] <= iHeaderRowIndex
    ]
    if not listPreviousCenters:
        raise ValueError(
            "店舗グループより上に配送センター見出しがありません。行 = "
            + str(iHeaderRowIndex + 1)
            + "、列 = "
            + str(iStoreCodeColumnIndex + 1)
        )
    iNearestCenterRowIndex: int = max(
        tupleCenter[0] for tupleCenter in listPreviousCenters
    )
    listSameRowCenters: list[tuple[int, int, str]] = sorted(
        (
            tupleCenter
            for tupleCenter in listPreviousCenters
            if tupleCenter[0] == iNearestCenterRowIndex
            and tupleCenter[1] <= iStoreCodeColumnIndex
        ),
        key=lambda tupleCenter: tupleCenter[1],
    )
    if not listSameRowCenters:
        raise ValueError(
            "店舗グループを配送センターへ関連付けできません。行 = "
            + str(iHeaderRowIndex + 1)
            + "、列 = "
            + str(iStoreCodeColumnIndex + 1)
        )
    return listSameRowCenters[-1]


def extract_store_group_rows(
    listRows: list[list[str]],
    tupleStoreGroup: tuple[int, int, int],
    pszDistributionCenterName: str,
) -> list[list[str]]:
    """店舗グループのヘッダー直後から小計まで店舗コード・店舗名を返します。"""
    iHeaderRowIndex, iStoreCodeColumnIndex, iStoreNameColumnIndex = tupleStoreGroup
    listMappingRows: list[list[str]] = []
    bFoundSubtotal: bool = False
    for listValues in listRows[iHeaderRowIndex + 1 :]:
        pszStoreCode: str = get_row_value(listValues, iStoreCodeColumnIndex)
        pszStoreName: str = get_row_value(listValues, iStoreNameColumnIndex)
        if pszStoreCode.strip() == "小計":
            bFoundSubtotal = True
            break
        if pszStoreCode.strip() == "" or pszStoreName.strip() == "":
            continue
        if pszStoreCode == "店舗コード" or pszStoreName.endswith("店舗名"):
            continue
        listMappingRows.append(
            [pszDistributionCenterName, pszStoreCode, pszStoreName]
        )
    if not bFoundSubtotal:
        raise ValueError(
            "店舗グループの小計行が見つかりません。配送センター = "
            + pszDistributionCenterName
            + "、ヘッダー行 = "
            + str(iHeaderRowIndex + 1)
            + "、店舗コード列 = "
            + str(iStoreCodeColumnIndex + 1)
        )
    if not listMappingRows:
        raise ValueError(
            "店舗グループに店舗データが見つかりません。配送センター = "
            + pszDistributionCenterName
            + "、ヘッダー行 = "
            + str(iHeaderRowIndex + 1)
            + "、店舗コード列 = "
            + str(iStoreCodeColumnIndex + 1)
        )
    return listMappingRows


def build_area_store_mapping_rows(
    listRows: list[list[str]],
) -> tuple[list[list[str]], int, int]:
    """配送センターと全店舗グループを検出して3列の対応表を返します。"""
    listCenters: list[tuple[int, int, str]] = find_distribution_centers(listRows)
    listStoreGroups: list[tuple[int, int, int]] = find_store_groups(listRows)
    listAssignedGroups: list[
        tuple[tuple[int, int, str], tuple[int, int, int]]
    ] = [
        (assign_distribution_center(tupleStoreGroup, listCenters), tupleStoreGroup)
        for tupleStoreGroup in listStoreGroups
    ]
    listAssignedGroups.sort(
        key=lambda tupleAssignment: (
            tupleAssignment[0][0],
            tupleAssignment[0][1],
            tupleAssignment[1][1],
            tupleAssignment[1][0],
        )
    )
    listMappingRows: list[list[str]] = []
    for tupleCenter, tupleStoreGroup in listAssignedGroups:
        listMappingRows.extend(
            extract_store_group_rows(
                listRows,
                tupleStoreGroup,
                tupleCenter[2],
            )
        )
    if not listMappingRows:
        raise ValueError("店舗データが見つかりません。")
    iCenterCount: int = len(
        {(tupleCenter[0], tupleCenter[1]) for tupleCenter, _ in listAssignedGroups}
    )
    return listMappingRows, iCenterCount, len(listAssignedGroups)


def save_area_store_mapping_tsv(
    objOutputPath: Path, listMappingRows: list[list[str]]
) -> None:
    """3列ヘッダーと週間店舗対応データをUTF-8 TSVへ保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow(AREA_STORE_MAPPING_HEADERS)
        objWriter.writerows(listMappingRows)


def get_area_store_mapping_output_path(objInputTsvPath: Path) -> Path:
    """週間TSVと同じフォルダーに作る店舗対応表TSVのパスを返します。"""
    return objInputTsvPath.with_name(AREA_STORE_MAPPING_FILE_NAME)


def process_area_store_mapping_file(
    objInputTsvPath: Path,
) -> tuple[Path, int, int, int]:
    """本州マグロ週間TSVから店舗対応表を独立して作成します。"""
    if not objInputTsvPath.exists() or not objInputTsvPath.is_file():
        raise ValueError("週間TSVが見つかりません。Path = " + str(objInputTsvPath))
    listRows: list[list[str]] = read_tsv_rows(objInputTsvPath)
    try:
        listMappingRows, iCenterCount, iGroupCount = build_area_store_mapping_rows(
            listRows
        )
    except ValueError as objException:
        raise ValueError(
            str(objException) + " Path = " + str(objInputTsvPath)
        ) from objException
    objOutputPath: Path = get_area_store_mapping_output_path(objInputTsvPath)
    objTemporaryPath: Path = create_temporary_path(objOutputPath)
    try:
        save_area_store_mapping_tsv(objTemporaryPath, listMappingRows)
        os.replace(objTemporaryPath, objOutputPath)
    finally:
        if objTemporaryPath.exists():
            objTemporaryPath.unlink()
    return objOutputPath, len(listMappingRows), iCenterCount, iGroupCount


def find_backup_numbers(objOutputPath: Path) -> list[int]:
    """通常名TSVに対応する既存バックアップの4桁番号を返します。"""
    objPattern: re.Pattern[str] = re.compile(
        r"^" + re.escape(objOutputPath.name) + r"\.bk([0-9]{4})\.tsv$"
    )
    listBackupNumbers: list[int] = []
    for objCandidatePath in objOutputPath.parent.iterdir():
        objMatch: re.Match[str] | None = objPattern.fullmatch(objCandidatePath.name)
        if objMatch is None or not objCandidatePath.is_file():
            continue
        iBackupNumber: int = int(objMatch.group(1))
        if 1 <= iBackupNumber <= MAX_BACKUP_NUMBER:
            listBackupNumbers.append(iBackupNumber)
    return listBackupNumbers


def get_next_backup_path(objOutputPath: Path) -> Path:
    """既存最大番号の次となる.bk%04d.tsvパスを返します。"""
    listBackupNumbers: list[int] = find_backup_numbers(objOutputPath)
    iBackupNumber: int = 1 if not listBackupNumbers else max(listBackupNumbers) + 1
    if iBackupNumber > MAX_BACKUP_NUMBER:
        raise ValueError(
            "バックアップ番号が最大値9999に到達しています。Path = "
            + str(objOutputPath)
        )
    return objOutputPath.with_name(
        objOutputPath.name + f".bk{iBackupNumber:04d}.tsv"
    )


def rename_output_to_backup(objOutputPath: Path) -> Path | None:
    """通常名TSVがあれば次の連番バックアップ名へ変更します。"""
    if not objOutputPath.exists():
        return None
    objBackupPath: Path = get_next_backup_path(objOutputPath)
    if objBackupPath.exists():
        raise FileExistsError(
            "バックアップ先がすでに存在します。Path = " + str(objBackupPath)
        )
    os.rename(objOutputPath, objBackupPath)
    return objBackupPath


def build_warning_text(
    pszInputFileFullPath: str,
    listMissingWorksheetNames: list[str],
    dictOutputPaths: dict[str, Path],
    dictBackupPaths: dict[str, Path],
    dictMappingOutputPaths: dict[str, Path],
    dictMappingBackupPaths: dict[str, Path],
) -> str:
    """未検出シート、作成TSV、旧TSVバックアップを含む警告文を返します。"""
    listLines: list[str] = [
        "処理結果: 警告",
        "入力ファイル: " + os.path.abspath(pszInputFileFullPath),
        "未検出シート: " + ", ".join(listMissingWorksheetNames),
        "警告内容: 対象シートが見つからないため、このシートのTSVは作成しませんでした。",
    ]
    for pszWorksheetName in TARGET_WORKSHEET_NAMES:
        if pszWorksheetName in listMissingWorksheetNames:
            continue
        listLines.append("作成したTSV: " + str(dictOutputPaths[pszWorksheetName]))
    for pszWorksheetName in listMissingWorksheetNames:
        if pszWorksheetName not in dictBackupPaths:
            continue
        listLines.append(
            "旧TSVの変更前パス: " + str(dictOutputPaths[pszWorksheetName])
        )
        listLines.append(
            "旧TSVのバックアップパス: " + str(dictBackupPaths[pszWorksheetName])
        )
    for pszWorksheetName in TARGET_WORKSHEET_NAMES:
        objMappingOutputPath: Path = dictMappingOutputPaths[pszWorksheetName]
        if pszWorksheetName in listMissingWorksheetNames:
            listLines.append(
                pszWorksheetName
                + "対応表: 対象シートがないため作成しませんでした。"
            )
            if pszWorksheetName in dictMappingBackupPaths:
                listLines.append(
                    "旧対応表の変更前パス: " + str(objMappingOutputPath)
                )
                listLines.append(
                    "旧対応表のバックアップパス: "
                    + str(dictMappingBackupPaths[pszWorksheetName])
                )
        else:
            listLines.append("作成予定の対応表: " + str(objMappingOutputPath))
    return "\n".join(listLines) + "\n"


def replace_output_files(
    dictTemporaryPaths: dict[Path, Path],
    dictMissingOutputBackups: dict[Path, Path],
    objWarningPath: Path,
    objTemporaryWarningPath: Path | None,
) -> None:
    """TSV・警告を置換し、欠落シートの旧TSVを連番名へ変更します。"""
    setManagedPaths: set[Path] = set(dictTemporaryPaths.keys()) | {objWarningPath}
    dictRollbackPaths: dict[Path, Path] = {}
    listRenamedMissingOutputs: list[tuple[Path, Path]] = []
    try:
        for objManagedPath in setManagedPaths:
            if not objManagedPath.exists():
                continue
            objRollbackPath: Path = create_temporary_path(objManagedPath)
            shutil.copy2(objManagedPath, objRollbackPath)
            dictRollbackPaths[objManagedPath] = objRollbackPath
        for objOutputPath, objBackupPath in dictMissingOutputBackups.items():
            if objBackupPath.exists():
                raise FileExistsError(
                    "バックアップ先がすでに存在します。Path = " + str(objBackupPath)
                )
            os.rename(objOutputPath, objBackupPath)
            listRenamedMissingOutputs.append((objOutputPath, objBackupPath))
        for objOutputPath, objTemporaryPath in dictTemporaryPaths.items():
            os.replace(objTemporaryPath, objOutputPath)
        if objTemporaryWarningPath is None:
            if objWarningPath.exists():
                objWarningPath.unlink()
        else:
            os.replace(objTemporaryWarningPath, objWarningPath)
    except Exception:
        for objManagedPath in setManagedPaths:
            objRollbackPath = dictRollbackPaths.get(objManagedPath)
            if objRollbackPath is not None and objRollbackPath.exists():
                os.replace(objRollbackPath, objManagedPath)
            elif objManagedPath.exists():
                objManagedPath.unlink()
        for objOutputPath, objBackupPath in reversed(listRenamedMissingOutputs):
            if objBackupPath.exists():
                os.rename(objBackupPath, objOutputPath)
        raise
    finally:
        for objRollbackPath in dictRollbackPaths.values():
            if objRollbackPath.exists():
                objRollbackPath.unlink()


def process_input_file(pszInputFileFullPath: str) -> None:
    """存在する対象Excelシートのセル値から調査用TSVを作成します。"""
    pszValidatedPath: str = validate_input_path(pszInputFileFullPath)
    dictWorksheetResults, listMissingWorksheetNames = read_excel_worksheets(
        pszValidatedPath
    )
    dictOutputPaths: dict[str, Path] = {
        pszWorksheetName: get_output_file_path(pszValidatedPath, pszWorksheetName)
        for pszWorksheetName in TARGET_WORKSHEET_NAMES
    }
    dictMappingOutputPaths: dict[str, Path] = {
        "割り": get_allocation_mapping_output_path(dictOutputPaths["割り"]),
        "本州マグロ(週間)": get_area_store_mapping_output_path(
            dictOutputPaths["本州マグロ(週間)"]
        ),
    }
    dictBackupPaths: dict[str, Path] = {}
    for pszWorksheetName in listMissingWorksheetNames:
        objMissingOutputPath: Path = dictOutputPaths[pszWorksheetName]
        if objMissingOutputPath.exists():
            dictBackupPaths[pszWorksheetName] = get_next_backup_path(
                objMissingOutputPath
            )
    dictMappingBackupPaths: dict[str, Path] = {}
    for pszWorksheetName in listMissingWorksheetNames:
        objMappingOutputPath: Path = dictMappingOutputPaths[pszWorksheetName]
        if objMappingOutputPath.exists():
            dictMappingBackupPaths[pszWorksheetName] = get_next_backup_path(
                objMappingOutputPath
            )
    dictTemporaryPaths: dict[Path, Path] = {}
    objWarningPath: Path = Path(get_warning_file_full_path(pszValidatedPath))
    objTemporaryWarningPath: Path | None = None
    try:
        for pszWorksheetName, (listRows, _) in dictWorksheetResults.items():
            objOutputPath: Path = dictOutputPaths[pszWorksheetName]
            objTemporaryPath: Path = create_temporary_path(objOutputPath)
            dictTemporaryPaths[objOutputPath] = objTemporaryPath
            save_tsv_rows(objTemporaryPath, listRows)
        if dictWorksheetResults and listMissingWorksheetNames:
            objTemporaryWarningPath = create_temporary_path(objWarningPath)
            write_error_text(
                str(objTemporaryWarningPath),
                build_warning_text(
                    pszValidatedPath,
                    listMissingWorksheetNames,
                    dictOutputPaths,
                    dictBackupPaths,
                    dictMappingOutputPaths,
                    dictMappingBackupPaths,
                ),
            )
        dictMissingOutputBackups: dict[Path, Path] = {
            dictOutputPaths[pszWorksheetName]: objBackupPath
            for pszWorksheetName, objBackupPath in dictBackupPaths.items()
        }
        for pszWorksheetName, objMappingBackupPath in (
            dictMappingBackupPaths.items()
        ):
            dictMissingOutputBackups[
                dictMappingOutputPaths[pszWorksheetName]
            ] = objMappingBackupPath
        replace_output_files(
            dictTemporaryPaths,
            dictMissingOutputBackups,
            objWarningPath,
            objTemporaryWarningPath,
        )
    finally:
        for objTemporaryPath in dictTemporaryPaths.values():
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()
        if objTemporaryWarningPath is not None and objTemporaryWarningPath.exists():
            objTemporaryWarningPath.unlink()
    if not dictWorksheetResults:
        raise ValueError(
            "対象シートが見つかりません。対象シート = "
            + ", ".join(TARGET_WORKSHEET_NAMES)
        )
    objCreatedAllocationMappingPath: Path | None = None
    iAllocationMappingRowCount: int = 0
    objCreatedMappingPath: Path | None = None
    iMappingRowCount: int = 0
    iMappingCenterCount: int = 0
    iMappingGroupCount: int = 0
    listMappingResultLines: list[str] = []
    listMappingErrorLines: list[str] = []
    if "割り" in dictWorksheetResults:
        try:
            (
                objCreatedAllocationMappingPath,
                iAllocationMappingRowCount,
            ) = process_allocation_mapping_file(dictOutputPaths["割り"])
            listMappingResultLines.append(
                "処理A（割り対応表）: 成功\n出力ファイル: "
                + str(objCreatedAllocationMappingPath)
            )
        except Exception as objException:
            try:
                objBackupPath = rename_output_to_backup(
                    dictMappingOutputPaths["割り"]
                )
                pszBackupDetail: str = (
                    ""
                    if objBackupPath is None
                    else "\n旧出力バックアップ: " + str(objBackupPath)
                )
            except Exception as objBackupException:
                pszBackupDetail = (
                    "\n旧出力のバックアップにも失敗しました。Detail = "
                    + str(objBackupException)
                )
            listMappingErrorLines.append(
                "処理A（割り対応表）: エラー\nエラー内容: "
                + str(objException)
                + pszBackupDetail
            )
    else:
        listMappingResultLines.append("処理A（割り対応表）: スキップ")
    if "本州マグロ(週間)" in dictWorksheetResults:
        try:
            (
                objCreatedMappingPath,
                iMappingRowCount,
                iMappingCenterCount,
                iMappingGroupCount,
            ) = process_area_store_mapping_file(
                dictOutputPaths["本州マグロ(週間)"]
            )
            listMappingResultLines.append(
                "処理B（週間配送センター対応表）: 成功\n出力ファイル: "
                + str(objCreatedMappingPath)
            )
        except Exception as objException:
            try:
                objBackupPath = rename_output_to_backup(
                    dictMappingOutputPaths["本州マグロ(週間)"]
                )
                pszBackupDetail = (
                    ""
                    if objBackupPath is None
                    else "\n旧出力バックアップ: " + str(objBackupPath)
                )
            except Exception as objBackupException:
                pszBackupDetail = (
                    "\n旧出力のバックアップにも失敗しました。Detail = "
                    + str(objBackupException)
                )
            listMappingErrorLines.append(
                "処理B（週間配送センター対応表）: エラー\nエラー内容: "
                + str(objException)
                + pszBackupDetail
            )
    else:
        listMappingResultLines.append("処理B（週間配送センター対応表）: スキップ")
    if listMappingErrorLines:
        raise ValueError(
            "\n\n".join(listMappingErrorLines + listMappingResultLines)
        )
    remove_old_error_file(pszValidatedPath)
    print("朝日注文エリア店舗対応調査用TSVファイルを作成しました。")
    print("Input: " + pszValidatedPath)
    for pszWorksheetName in TARGET_WORKSHEET_NAMES:
        if pszWorksheetName not in dictWorksheetResults:
            continue
        listRows, iColumnCount = dictWorksheetResults[pszWorksheetName]
        print("Worksheet: " + pszWorksheetName)
        print("TSV: " + str(dictOutputPaths[pszWorksheetName]))
        print("Rows: " + str(len(listRows)))
        print("Columns: " + str(iColumnCount))
    for pszWorksheetName in listMissingWorksheetNames:
        print("Warning: 対象シートが見つかりません。Sheet = " + pszWorksheetName)
        if pszWorksheetName in dictBackupPaths:
            print("Backup: " + str(dictBackupPaths[pszWorksheetName]))
    if listMissingWorksheetNames:
        print("Warning File: " + str(objWarningPath))
    if objCreatedMappingPath is not None:
        print(
            "Weekly Area Store Mapping Input: "
            + str(dictOutputPaths["本州マグロ(週間)"])
        )
        print("Weekly Area Store Mapping TSV: " + str(objCreatedMappingPath))
        print("Weekly Area Store Mapping Rows: " + str(iMappingRowCount))
        print("Weekly Area Store Mapping Centers: " + str(iMappingCenterCount))
        print("Weekly Area Store Mapping Groups: " + str(iMappingGroupCount))
    if objCreatedAllocationMappingPath is not None:
        print("Allocation Area Store Mapping Result: Success")
        print("Allocation Area Store Mapping Input: " + str(dictOutputPaths["割り"]))
        print(
            "Allocation Area Store Mapping TSV: "
            + str(objCreatedAllocationMappingPath)
        )
        print(
            "Allocation Area Store Mapping Rows: "
            + str(iAllocationMappingRowCount)
        )


def main() -> int:
    """引数を確認して処理し、成功0・失敗1の終了コードを返します。"""
    if len(sys.argv) != 2:
        pszScriptFileName: str = os.path.basename(__file__)
        pszErrorMessage: str = (
            "Error: 入力Excelファイルパスを1件指定してください。\n"
            + "Usage: python "
            + pszScriptFileName
            + " <input_xlsx_file_path>\n"
        )
        print(pszErrorMessage, file=sys.stderr, end="")
        pszErrorFileFullPath: str = (
            os.path.splitext(pszScriptFileName)[0] + "_error_argument.txt"
        )
        try:
            write_error_text(pszErrorFileFullPath, pszErrorMessage)
        except OSError as objException:
            print(
                "Error: 引数エラーファイルを保存できません。Detail = "
                + str(objException),
                file=sys.stderr,
            )
        return 1
    pszInputFileFullPath: str = sys.argv[1]
    try:
        process_input_file(pszInputFileFullPath)
    except Exception as objException:
        report_processing_error(
            pszInputFileFullPath,
            "朝日注文エリア店舗対応調査TSV作成処理",
            str(objException),
        )
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
